#!/usr/bin/env python3
"""
读取 meetings/special/fu/原始数据/tiktok0427.xlsx
生成主播数据报告 JSON + HTML
"""

import openpyxl
from datetime import datetime, timedelta
import json
import os

EXCEL_PATH = "meetings/special/fu/原始数据/tiktok0427.xlsx"
OUTPUT_HTML = "src/report/streamer-data.html"
OUTPUT_JSON = "src/report/streamer-data.json"

# Excel epoch for Windows (1899-12-30)
EXCEL_EPOCH = datetime(1899, 12, 30)

def excel_date_to_datetime(serial):
    if serial is None:
        return None
    try:
        return EXCEL_EPOCH + timedelta(days=int(serial))
    except Exception:
        return None

def parse_time_str(time_val):
    """提取小时数，用于判断下午场/晚间场"""
    if time_val is None:
        return None
    if isinstance(time_val, (int, float)):
        h = int(time_val)
        # 简单启发：如果数字<=5 可能是下午（16-17点），但需要结合主播判断
        return h
    s = str(time_val).strip()
    # 匹配 "下午1:55", "下午9:09", "上午9:14", "下午11"
    if "上午" in s or "早上" in s or "凌晨" in s:
        m = re.search(r'(\d+)', s)
        if m:
            return int(m.group(1))
    elif "下午" in s or "晚上" in s:
        m = re.search(r'(\d+)', s)
        if m:
            h = int(m.group(1))
            return h + 12 if h < 12 else h
    elif "中午" in s:
        return 12
    else:
        m = re.search(r'(\d+)', s)
        if m:
            return int(m.group(1))
    return None

def is_afternoon(hour):
    return 12 <= hour < 18

def is_evening(hour):
    return 18 <= hour < 24

import re

def read_sheet(ws, sheet_name):
    """读取一个 sheet，返回主播信息和每日数据列表"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return None
    
    # 汇总信息在第1行 (index 0)
    # 总粉丝量在第2行 (index 1)
    # 表头在第5行 (index 4)
    # 数据从第6行 (index 5)
    header_row = rows[4]
    
    # 找到各列索引
    # 列: B=日期, C=直播时间, D=直播时长, E=点赞, F=评论, G=粉丝平均观看时长,
    #     H=涨粉, I=观看次数, J=尖峰在线, K=平均观看时长, L=钻石收入, M=人民币收入
    # openpyxl values_only 返回的是 tuple，索引从0开始对应 A,B,C...
    # 但实际上前面的打印显示 B=index1 是 None, C=index2 是 'data', D=index3 是 'time'
    # 第5行: (None, None, '发布日期', '直播时间', '直播时长', '点赞量', '评论量', ...)
    # 所以 index2=发布日期, index3=直播时间, index4=直播时长, index5=点赞量, index6=评论量,
    # index7=粉丝平均观看时长, index8=涨粉, index9=观看次数, index10=尖峰在线, index11=平均观看时长,
    # index12=钻石收入, index13=人民币收入
    
    daily = []
    for row in rows[5:]:
        if row[2] is None:
            continue
        dt = excel_date_to_datetime(row[2])
        if dt is None:
            continue
        
        time_val = row[3]
        hour = parse_time_str(time_val)
        
        rmb_income = 0
        if row[13] is not None:
            try:
                rmb_income = float(row[13])
            except Exception:
                rmb_income = 0
        
        diamond = 0
        if row[12] is not None:
            try:
                diamond = float(row[12])
            except Exception:
                diamond = 0
        
        def to_float(v):
            if v is None:
                return 0
            try:
                return float(v)
            except Exception:
                return 0
        
        def to_int(v):
            if v is None:
                return 0
            try:
                return int(float(v))
            except Exception:
                return 0
        
        entry = {
            "date": dt.date(),
            "datetime": dt,
            "hour": hour,
            "time_str": str(time_val) if time_val else "",
            "likes": to_int(row[5]),
            "comments": to_int(row[6]),
            "fan_duration": to_int(row[7]),
            "followers": to_int(row[8]),
            "views": to_int(row[9]),
            "peak": to_int(row[10]),
            "avg_duration": to_int(row[11]),
            "diamond": diamond,
            "income": rmb_income,
        }
        daily.append(entry)
    
    # 汇总信息
    # 第0行 index6 是总粉丝量（直播粉丝增长），index10 是总收入钻石，index13 是约人民币
    # 实际上第0行: (None, None, '抖音', '志青Ricon', None, None, '直播粉丝增长', 34372, None, None, '总收入钻石', 274920, '约人民币', 7697.76, ...)
    # 总粉丝量在第1行 index3
    total_followers = 0
    if len(rows) > 1 and rows[1][3] is not None:
        try:
            total_followers = int(float(rows[1][3]))
        except Exception:
            pass
    
    return {
        "name": sheet_name,
        "total_followers": total_followers,
        "daily": daily
    }

def merge_daily(entries):
    """按日期合并同一天的直播场次"""
    by_date = {}
    for e in entries:
        d = e["date"]
        if d not in by_date:
            by_date[d] = []
        by_date[d].append(e)
    
    merged = []
    for d in sorted(by_date.keys()):
        items = by_date[d]
        total_income = sum(x["income"] for x in items)
        total_views = sum(x["views"] for x in items)
        total_likes = sum(x["likes"] for x in items)
        total_comments = sum(x["comments"] for x in items)
        total_followers = sum(x["followers"] for x in items)
        peak = max(x["peak"] for x in items) if items else 0
        
        # 时长取平均
        fan_durations = [x["fan_duration"] for x in items if x["fan_duration"] > 0]
        avg_fan_duration = sum(fan_durations) / len(fan_durations) if fan_durations else 0
        
        avg_durations = [x["avg_duration"] for x in items if x["avg_duration"] > 0]
        avg_duration = sum(avg_durations) / len(avg_durations) if avg_durations else 0
        
        merged.append({
            "date": d,
            "income": round(total_income, 2),
            "views": total_views,
            "likes": total_likes,
            "comments": total_comments,
            "followers": total_followers,
            "peak": peak,
            "fan_duration": round(avg_fan_duration, 1),
            "avg_duration": round(avg_duration, 1),
        })
    return merged

def get_last_n_days(all_dates, n):
    """获取最近n个有数据的日期（所有主播的并集）"""
    sorted_dates = sorted(list(set(all_dates)))
    return sorted_dates[-n:]

def moving_average(data, window=7):
    result = []
    for i in range(len(data)):
        start = max(0, i - window + 1)
        vals = [data[j] for j in range(start, i + 1) if data[j] is not None]
        if vals:
            result.append(round(sum(vals) / len(vals), 2))
        else:
            result.append(0)
    return result

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    streamers = {}
    all_dates = set()
    
    for sheet_name in wb.sheetnames:
        info = read_sheet(wb[sheet_name], sheet_name)
        if info is None:
            continue
        merged = merge_daily(info["daily"])
        streamers[sheet_name] = {
            "name": sheet_name,
            "total_followers": info["total_followers"],
            "daily": merged
        }
        for m in merged:
            all_dates.add(m["date"])
    
    # 主播名字映射
    name_map = {
        "Heni": "阿贤heni",
        "Ricon": "志青Ricon",
        "MYMY": "MYMY",
        "阿园Chua": "阿园Chua",
        "Pink": "Pink",
        "Peo": "peo",
        "Vanie": "Vanie",
    }
    
    # 最近30天和10天的日期范围
    dates_30 = get_last_n_days(all_dates, 30)
    dates_10 = dates_30[-10:]
    
    # 构建以日期为键的数据
    def build_series(streamer_key, field):
        s = streamers[streamer_key]
        by_date = {x["date"]: x[field] for x in s["daily"]}
        return [by_date.get(d, 0) for d in dates_30]
    
    def build_series_10(streamer_key, field):
        s = streamers[streamer_key]
        by_date = {x["date"]: x[field] for x in s["daily"]}
        return [by_date.get(d, 0) for d in dates_10]
    
    # 总收入/总观看序列
    total_income_30 = [0] * 30
    total_views_30 = [0] * 30
    for sk in streamers:
        inc = build_series(sk, "income")
        vw = build_series(sk, "views")
        for i in range(30):
            total_income_30[i] = round(total_income_30[i] + inc[i], 2)
            total_views_30[i] += vw[i]
    
    total_income_ma = moving_average(total_income_30, 7)
    total_views_ma = moving_average(total_views_30, 7)
    
    labels30 = [d.strftime("%Y-%m-%d") for d in dates_30]
    labels10 = [d.strftime("%Y-%m-%d") for d in dates_10]
    
    # 周对周对比：最近7天 vs 前7天
    week1_dates = dates_30[-7:]   # 近一周
    week2_dates = dates_30[-14:-7] # 上一周
    
    def sum_week(streamer_key, dates_list):
        s = streamers[streamer_key]
        by_date = {x["date"]: x["income"] for x in s["daily"]}
        return round(sum(by_date.get(d, 0) for d in dates_list), 2)
    
    def sum_views_week(streamer_key, dates_list):
        s = streamers[streamer_key]
        by_date = {x["date"]: x["views"] for x in s["daily"]}
        return sum(by_date.get(d, 0) for d in dates_list)
    
    wow_data = []
    for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
        prev = sum_week(sk, week2_dates)
        curr = sum_week(sk, week1_dates)
        diff = round(curr - prev, 2)
        pct = round((diff / prev * 100), 1) if prev > 0 else 0
        if abs(pct) < 10:
            trend = "→ 稳定"
        elif pct > 0:
            trend = "↑ 增长"
        else:
            trend = "↓ 下滑"
        wow_data.append({
            "key": sk,
            "name": name_map[sk],
            "prev": prev,
            "curr": curr,
            "diff": diff,
            "pct": pct,
            "trend": trend
        })
    
    total_week1_income = sum(x["curr"] for x in wow_data)
    total_week2_income = sum(x["prev"] for x in wow_data)
    total_week1_views = sum(sum_views_week(x["key"], week1_dates) for x in wow_data)
    total_week2_views = sum(sum_views_week(x["key"], week2_dates) for x in wow_data)
    
    income_pct = round((total_week1_income - total_week2_income) / total_week2_income * 100, 1) if total_week2_income > 0 else 0
    views_pct = round((total_week1_views - total_week2_views) / total_week2_views * 100, 1) if total_week2_views > 0 else 0
    
    # 双场数据：MYMY 和 阿贤(Heni)
    # 需要原始未合并的场次数据
    
    # 实际上 streamers 里存的是合并后的，需要重新读原始
    def get_raw_daily(streamer_key):
        ws = wb[streamer_key]
        info = read_sheet(ws, streamer_key)
        return info["daily"] if info else []
    
    def process_double(streamer_key):
        raw = get_raw_daily(streamer_key)
        by_date = {}
        for e in raw:
            d = e["date"]
            h = e["hour"]
            if d not in by_date:
                by_date[d] = {"afternoon_income": 0, "evening_income": 0, "afternoon_peak": 0, "evening_peak": 0}
            if h is None:
                continue
            if is_afternoon(h):
                by_date[d]["afternoon_income"] += e["income"]
                by_date[d]["afternoon_peak"] = max(by_date[d]["afternoon_peak"], e["peak"])
            elif is_evening(h):
                by_date[d]["evening_income"] += e["income"]
                by_date[d]["evening_peak"] = max(by_date[d]["evening_peak"], e["peak"])
        
        sorted_dates = sorted(by_date.keys())
        dates_fmt = [d.strftime("%m/%d") for d in sorted_dates]
        return {
            "dates": dates_fmt,
            "afternoon_income": [round(by_date[d]["afternoon_income"], 2) for d in sorted_dates],
            "evening_income": [round(by_date[d]["evening_income"], 2) for d in sorted_dates],
            "afternoon_peak": [by_date[d]["afternoon_peak"] for d in sorted_dates],
            "evening_peak": [by_date[d]["evening_peak"] for d in sorted_dates],
        }
    
    mymy_double = process_double("MYMY")
    heni_double = process_double("Heni")
    
    # 读取运营事件，过滤主播类事件，生成 Chart.js annotation 配置
    events_path = "data/exports/revenue_events.json"
    streamer_events = []
    if os.path.exists(events_path):
        with open(events_path, "r", encoding="utf-8") as f:
            events_data = json.load(f)
        for ev in events_data.get("events", []):
            if ev.get("category") == "主播":
                streamer_events.append(ev)
    
    date_to_income_idx = {labels30[i]: i for i in range(30)}
    annotations = {}
    for ev in streamer_events:
        ev_id = ev["name"].replace(" ", "_").replace("\u3000", "_")
        ev_type = ev.get("type", "markPoint")
        color = ev.get("color", "#fc8452")
        date_str = ev["date"]
        note = ev.get("note", "")
        base = {
            "note": note,
            "label": {
                "display": True,
                "content": ev["name"],
                "position": "start",
                "backgroundColor": color,
                "color": "white",
                "font": {"size": 10}
            }
        }
        if ev_type == "markLine":
            annotations[ev_id] = {
                **base,
                "type": "line",
                "xMin": date_str,
                "xMax": date_str,
                "borderColor": color,
                "borderWidth": 2,
                "borderDash": [6, 6],
            }
        elif ev_type == "markPoint":
            y_val = 0
            if date_str in date_to_income_idx:
                y_val = total_income_30[date_to_income_idx[date_str]]
            annotations[ev_id] = {
                **base,
                "type": "point",
                "xValue": date_str,
                "yValue": y_val,
                "backgroundColor": color,
                "radius": 6,
            }
        elif ev_type == "markArea":
            end_date = ev.get("endDate", date_str)
            annotations[ev_id] = {
                **base,
                "type": "box",
                "xMin": date_str,
                "xMax": end_date,
                "backgroundColor": color + "33",
                "borderWidth": 0,
            }
    
    # peo / Vanie 的独立数据（全部历史）
    def build_all_income(streamer_key):
        s = streamers[streamer_key]
        return [x["income"] for x in s["daily"]], [x["date"].strftime("%m/%d") for x in s["daily"]]
    
    peo_income_all, peo_labels_all = build_all_income("Peo")
    vanie_income_all, vanie_labels_all = build_all_income("Vanie")
    
    # 饼图数据：近一周
    pie_views = []
    pie_income = []
    pie_names = []
    for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
        pie_names.append(name_map[sk])
        pie_income.append(sum_week(sk, week1_dates))
        pie_views.append(sum_views_week(sk, week1_dates))
    
    # stack 数据
    stack_views = {}
    stack_income = {}
    for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
        nk = name_map[sk]
        stack_views[nk] = build_series_10(sk, "views")
        stack_income[nk] = build_series_10(sk, "income")
    
    # 连胜统计（单赛场规则：按日收入排名，缺席即断连胜）
    def calculate_streaks():
        """计算每位主播的连胜记录"""
        # 每日冠军
        daily_results = []
        for d in dates_30:
            winner = None
            max_income = -1
            for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
                s = streamers[sk]
                by_date = {x["date"]: x["income"] for x in s["daily"]}
                inc = by_date.get(d, 0)
                if inc > max_income:
                    max_income = inc
                    winner = name_map[sk]
            daily_results.append({"date": d, "winner": winner, "income": max_income})
        
        streaks = {}
        for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
            nm = name_map[sk]
            s = streamers[sk]
            by_date = {x["date"]: x["income"] for x in s["daily"]}
            
            current = 0
            max_streak = 0
            active = False
            
            for dr in daily_results:
                d = dr["date"]
                inc = by_date.get(d, 0)
                if inc == 0:
                    # 缺席，连胜中断
                    current = 0
                    active = False
                elif dr["winner"] == nm:
                    # 获胜，连胜+1
                    current += 1
                    active = True
                    max_streak = max(max_streak, current)
                else:
                    # 参赛但未获胜，连胜中断
                    current = 0
                    active = False
            
            streaks[nm] = {
                "max_streak": max_streak,
                "current_streak": current if active else 0
            }
        return streaks
    
    streak_data = calculate_streaks()
    
    # 构造 JSON 数据对象
    data = {
        "labels30": labels30,
        "labels10": labels10,
        "totalViewsData": total_views_30,
        "totalViewsMA": total_views_ma,
        "totalIncomeData": total_income_30,
        "totalIncomeMA": total_income_ma,
        "peoLabels": peo_labels_all,
        "peoIncomeData": peo_income_all,
        "vanieLabels": vanie_labels_all,
        "vanieIncomeData": vanie_income_all,
        "mymyDates": mymy_double["dates"],
        "mymyAfternoonIncome": mymy_double["afternoon_income"],
        "mymyEveningIncome": mymy_double["evening_income"],
        "mymyAfternoonPeak": mymy_double["afternoon_peak"],
        "mymyEveningPeak": mymy_double["evening_peak"],
        "heniDates": heni_double["dates"],
        "heniAfternoonIncome": heni_double["afternoon_income"],
        "heniEveningIncome": heni_double["evening_income"],
        "heniAfternoonPeak": heni_double["afternoon_peak"],
        "heniEveningPeak": heni_double["evening_peak"],
        "pieLabels": pie_names,
        "pieViewsData": pie_views,
        "pieIncomeData": pie_income,
        "stackLabels": labels10,
        "stackViewsData": stack_views,
        "stackIncomeData": stack_income,
        "annotations": annotations,
        "streakData": streak_data,
    }
    
    # 各主播 30天 序列
    for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
        nk = name_map[sk]
        data[f"{nk}_income"] = build_series(sk, "income")
        data[f"{nk}_peak"] = build_series(sk, "peak")
        data[f"{nk}_followers"] = build_series(sk, "followers")
        data[f"{nk}_likes"] = build_series(sk, "likes")
        data[f"{nk}_interaction"] = [a + b for a, b in zip(build_series(sk, "likes"), build_series(sk, "comments"))]
        data[f"{nk}_fanDuration"] = build_series(sk, "fan_duration")
        data[f"{nk}_avgDuration"] = build_series(sk, "avg_duration")
        data[f"{nk}_views"] = build_series(sk, "views")
        data[f"{nk}_income_10"] = build_series_10(sk, "income")
        data[f"{nk}_views_10"] = build_series_10(sk, "views")
    
    # 写入 JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    # 生成 HTML
    # 读取现有模板
    with open(OUTPUT_HTML, "r", encoding="utf-8") as f:
        template = f.read()
    
    # Replace embedded JSON data object with latest data (including annotations)
    json_str = json.dumps(data, ensure_ascii=False)
    start_marker = 'const data = '
    end_marker = ';\n\nfunction showTab'
    s = template.find(start_marker)
    e = template.find(end_marker, s)
    if s != -1 and e != -1:
        template = template[:s + len(start_marker)] + json_str + template[e:]
    else:
        # fallback to regex
        template = re.sub(
            r'const data = \{.*?\};\s*renderCharts\(data\);',
            f'const data = {json_str};\n        renderCharts(data);',
            template,
            flags=re.DOTALL
        )
    
    # 更新顶部信息
    start_date = dates_30[0].strftime("%Y-%m-%d")
    end_date = dates_30[-1].strftime("%Y-%m-%d")
    update_date = end_date
    
    # 替换 header 信息
    template = re.sub(
        r'<p>数据更新时间: .*?</p>',
        f'<p>数据更新时间: {update_date} | 分析周期: {start_date} 至 {end_date}</p>',
        template
    )
    
    # 替换统计卡片
    template = re.sub(
        r'<div class="stat-value">¥[\d,\.]+</div>\s*<div class="stat-label">近一周总收入</div>',
        f'<div class="stat-value">¥{total_week1_income:,.0f}</div>\n                <div class="stat-label">近一周总收入</div>',
        template
    )
    template = re.sub(
        r'<div class="stat-value">[\d\.]+万</div>\s*<div class="stat-label">近一周总观看</div>',
        f'<div class="stat-value">{total_week1_views/10000:.1f}万</div>\n                <div class="stat-label">近一周总观看</div>',
        template
    )
    
    views_pct_str = f"+{views_pct}%" if views_pct >= 0 else f"{views_pct}%"
    income_pct_str = f"+{income_pct}%" if income_pct >= 0 else f"{income_pct}%"
    
    template = re.sub(
        r'<div class="stat-value">[\+\-]?\d+\.?\d*%</div>\s*<div class="stat-label">观看周环比</div>',
        f'<div class="stat-value">{views_pct_str}</div>\n                <div class="stat-label">观看周环比</div>',
        template
    )
    template = re.sub(
        r'<div class="stat-value">[\+\-]?\d+\.?\d*%</div>\s*<div class="stat-label">收入周环比</div>',
        f'<div class="stat-value">{income_pct_str}</div>\n                <div class="stat-label">收入周环比</div>',
        template
    )
    
    # 更新关键发现文字
    w2_start = week2_dates[0].strftime("%m/%d")
    w2_end = week2_dates[-1].strftime("%m/%d")
    w1_start = week1_dates[0].strftime("%m/%d")
    w1_end = week1_dates[-1].strftime("%m/%d")
    
    trend_text = f"""
                    <strong>关键发现：</strong>{w1_start}-{w1_end} vs {w2_start}-{w2_end}，
                    总观看次数{"增长" if views_pct >= 0 else "下降"}<strong>{views_pct_str}</strong>，
                    总收入{"增长" if income_pct >= 0 else "下降"}<strong>{income_pct_str}</strong>。
    """.strip()
    
    template = re.sub(
        r'<p style="margin-top:15px;color:#666">\s*<strong>关键发现：</strong>.*?</p>',
        f'<p style="margin-top:15px;color:#666">\n                    {trend_text}\n                </p>',
        template
    )
    
    # 更新周对周表格
    table_rows = []
    for row in wow_data:
        diff_class = "positive" if row["diff"] >= 0 else "negative"
        pct_str = f"+{row['pct']}%" if row["pct"] >= 0 else f"{row['pct']}%"
        table_rows.append(
            f'<tr><td><strong>{row["name"]}</strong></td><td>¥{row["prev"]}</td><td>¥{row["curr"]}</td>'
            f'<td class="{diff_class}">{row["diff"]:+.1f}</td><td class="{diff_class}">{pct_str}</td><td>{row["trend"]}</td></tr>'
        )
    table_html = "\n            ".join(table_rows)
    
    template = re.sub(
        r'<table>\s*<tr><th>主播</th><th>上周</th><th>本周</th><th>变化额</th><th>变化率</th><th>趋势</th></tr>.*?</table>',
        f'<table>\n                    <tr><th>主播</th><th>上周</th><th>本周</th><th>变化额</th><th>变化率</th><th>趋势</th></tr>\n                    {table_html}\n                </table>',
        template,
        flags=re.DOTALL
    )
    
    # 更新连胜统计模块
    streak_cards = []
    for sk in ["Heni", "MYMY", "Ricon", "阿园Chua", "Pink", "Peo", "Vanie"]:
        nm = name_map[sk]
        info = streak_data[nm]
        active_cls = ' active' if info['current_streak'] > 0 else ''
        streak_cards.append(
            f'<div class="streak-card{active_cls}">'
            f'<div class="streak-name">{nm}</div>'
            f'<div class="streak-value">{info["current_streak"]}</div>'
            f'<div class="streak-label">当前连胜</div>'
            f'<div class="streak-max">最高 {info["max_streak"]} 连胜</div>'
            f'</div>'
        )
    streak_html = '\n                '.join(streak_cards)
    
    template = re.sub(
        r'(<div class="streak-grid" id="streakGrid">)(.*?)(</div>\s*<div class="info")',
        f'\\1\n                {streak_html}\n            \\3',
        template,
        flags=re.DOTALL
    )
    
    # 更新双场对比统计
    # MYMY
    mymy_afternoon_avg = round(sum(mymy_double["afternoon_income"]) / len([x for x in mymy_double["afternoon_income"] if x > 0]), 0) if any(x > 0 for x in mymy_double["afternoon_income"]) else 0
    mymy_evening_avg = round(sum(mymy_double["evening_income"]) / len([x for x in mymy_double["evening_income"] if x > 0]), 0) if any(x > 0 for x in mymy_double["evening_income"]) else 0
    
    template = re.sub(
        r'<div class="comparison-label">下午场平均收入</div>\s*<div class="comparison-value">¥\d+</div>',
        f'<div class="comparison-label">下午场平均收入</div>\n                        <div class="comparison-value">¥{int(mymy_afternoon_avg)}</div>',
        template,
        count=1
    )
    template = re.sub(
        r'<div class="comparison-label">晚间场平均收入</div>\s*<div class="comparisonValue">¥\d+</div>',
        f'<div class="comparison-label">晚间场平均收入</div>\n                        <div class="comparison-value">¥{int(mymy_evening_avg)}</div>',
        template,
        count=1
    )
    
    # 阿贤
    heni_afternoon_avg = round(sum(heni_double["afternoon_income"]) / len([x for x in heni_double["afternoon_income"] if x > 0]), 0) if any(x > 0 for x in heni_double["afternoon_income"]) else 0
    heni_evening_avg = round(sum(heni_double["evening_income"]) / len([x for x in heni_double["evening_income"] if x > 0]), 0) if any(x > 0 for x in heni_double["evening_income"]) else 0
    
    # 替换第二组（阿贤）的双场数据
    # 使用更精确的正则
    heni_section = re.search(
        r'(阿贤heni 双场收入对比.*?<div class="comparison-box">.*?<div class="comparison-item">.*?<div class="comparison-label">下午场平均收入</div>\s*)<div class="comparison-value">¥\d+</div>',
        template,
        re.DOTALL
    )
    if heni_section:
        template = template[:heni_section.end()] + f'<div class="comparison-value">¥{int(heni_afternoon_avg)}</div>' + template[heni_section.end():]
    
    heni_section2 = re.search(
        r'(阿贤heni 双场收入对比.*?<div class="comparison-box">.*?<div class="comparison-item">.*?<div class="comparison-label">晚间场平均收入</div>\s*)<div class="comparison-value">¥\d+</div>',
        template,
        re.DOTALL
    )
    if heni_section2:
        template = template[:heni_section2.end()] + f'<div class="comparison-value">¥{int(heni_evening_avg)}</div>' + template[heni_section2.end():]
    
    # 修复 mymy 的 comparisonValue typo（老数据中是 comparisonValue 而非 comparison-value）
    template = template.replace('class="comparisonValue"', 'class="comparison-value"')
    

    # No external annotation CDN needed; using custom pure-Chart.js plugin
    
    # Inject custom annotation plugin before renderCharts
    # 模板已内置带 tooltip 的 annotation plugin，跳过注入
    rc_marker = '__PLUGIN_ALREADY_IN_TEMPLATE__'
    plugin_code = """        // Custom event annotation plugin (pure Chart.js, no external lib)
        const eventAnnotations = {
            id: 'eventAnnotations',
            _store: {},
            afterDraw(chart, args, options) {
                const ctx = chart.ctx;
                const xAxis = chart.scales.x || chart.scales['x'];
                const yAxis = chart.scales.y || chart.scales['y'];
                if (!xAxis || !yAxis) return;
                const annotations = eventAnnotations._store[chart.canvas.id] || {};
                const ca = chart.chartArea;
                if (!ca) return;
                for (const key in annotations) {
                    const ann = annotations[key];
                    const x = xAxis.getPixelForValue(ann.xMin || ann.xValue || ann.date);
                    if (x === undefined || x === null || isNaN(x)) continue;
                    ctx.save();
                    ctx.beginPath();
                    ctx.moveTo(x, ca.top);
                    ctx.lineTo(x, ca.bottom);
                    ctx.strokeStyle = ann.borderColor || ann.backgroundColor || '#e74c3c';
                    ctx.lineWidth = ann.borderWidth || 2;
                    ctx.setLineDash(ann.borderDash || [6, 6]);
                    ctx.stroke();
                    if (ann.label && ann.label.display) {
                        const text = ann.label.content;
                        const fontSize = (ann.label.font && ann.label.font.size) || 12;
                        ctx.font = `bold ${fontSize}px sans-serif`;
                        const pad = 6;
                        const m = ctx.measureText(text);
                        const w = m.width + pad * 2;
                        const h = fontSize + pad * 2;
                        let lx = x + 8;
                        let ly = ca.top + 8;
                        if (lx + w > ca.right) lx = x - w - 8;
                        ctx.fillStyle = ann.label.backgroundColor || ann.borderColor || '#e74c3c';
                        ctx.fillRect(lx, ly, w, h);
                        ctx.fillStyle = ann.label.color || '#fff';
                        ctx.textBaseline = 'middle';
                        ctx.fillText(text, lx + pad, ly + h / 2);
                    }
                    ctx.restore();
                }
            }
        };
        Chart.register(eventAnnotations);
        """
    if rc_marker in template:
        template = template.replace(rc_marker, plugin_code + rc_marker)
    
    # 修改 createLineChart 支持 annotations
    old_func = """        function createLineChart(ctx, labels, datasets, title) {
            return new Chart(ctx, {
                type: 'line',
                data: { labels, datasets },
                options: {
                    responsive: true, maintainAspectRatio: false,
                    interaction: { mode: 'index', intersect: false },
                    plugins: { legend: { position: 'top' }, title: { display: !!title, text: title } },
                    scales: { y: { beginAtZero: true } }
                }
            });
        }"""
    new_func = """        function createLineChart(ctx, labels, datasets, title, annotationOptions) {
            const plugins = { legend: { position: 'top' }, title: { display: !!title, text: title } };
            if (annotationOptions) {
                plugins.annotation = annotationOptions;
            }
            return new Chart(ctx, {
                type: 'line',
                data: { labels, datasets },
                options: {
                    responsive: true, maintainAspectRatio: false,
                    interaction: { mode: 'index', intersect: false },
                    plugins: plugins,
                    scales: { y: { beginAtZero: true } }
                }
            });
        }"""
    template = template.replace(old_func, new_func)
    
    # 在 totalIncomeChart 和 incomeChart 上应用 annotations
    old_total = """createLineChart(document.getElementById('totalIncomeChart').getContext('2d'), labels30, [
            { label: '总收入', data: totalIncomeData, borderColor: '#27ae60', backgroundColor: 'rgba(39,174,96,0.1)', tension: 0.4, fill: true },
            { label: '7日移动平均', data: totalIncomeMA, borderColor: '#e74c3c', borderDash: [5,5], tension: 0.4, pointRadius: 0 }
        ]);"""
    new_total = """createLineChart(document.getElementById('totalIncomeChart').getContext('2d'), labels30, [
            { label: '总收入', data: totalIncomeData, borderColor: '#27ae60', backgroundColor: 'rgba(39,174,96,0.1)', tension: 0.4, fill: true },
            { label: '7日移动平均', data: totalIncomeMA, borderColor: '#e74c3c', borderDash: [5,5], tension: 0.4, pointRadius: 0 }
        ], null, data.annotations || {});"""
    template = template.replace(old_total, new_total)
    
    old_income_chart = """createLineChart(document.getElementById('incomeChart').getContext('2d'), labels30, [
            { label: '阿贤', data: 阿贤heni_income, borderColor: '#e74c3c', tension: 0.3 },
            { label: 'MYMY', data: MYMY_income, borderColor: '#2ecc71', tension: 0.3 },
            { label: '志青', data: 志青Ricon_income, borderColor: '#95a5a6', tension: 0.3 },
            { label: '阿园', data: 阿园Chua_income, borderColor: '#f39c12', tension: 0.3 },
            { label: 'Pink', data: Pink_income, borderColor: '#9b59b6', tension: 0.3 },
            { label: 'peo', data: peo_income, borderColor: '#1abc9c', tension: 0.3 },
            { label: 'Vanie', data: Vanie_income, borderColor: '#3498db', tension: 0.3 }
        ]);"""
    new_income_chart = """createLineChart(document.getElementById('incomeChart').getContext('2d'), labels30, [
            { label: '阿贤', data: 阿贤heni_income, borderColor: '#e74c3c', tension: 0.3 },
            { label: 'MYMY', data: MYMY_income, borderColor: '#2ecc71', tension: 0.3 },
            { label: '志青', data: 志青Ricon_income, borderColor: '#95a5a6', tension: 0.3 },
            { label: '阿园', data: 阿园Chua_income, borderColor: '#f39c12', tension: 0.3 },
            { label: 'Pink', data: Pink_income, borderColor: '#9b59b6', tension: 0.3 },
            { label: 'peo', data: peo_income, borderColor: '#1abc9c', tension: 0.3 },
            { label: 'Vanie', data: Vanie_income, borderColor: '#3498db', tension: 0.3 }
        ], null, data.annotations || {});"""
    template = template.replace(old_income_chart, new_income_chart)

        # 写入 HTML
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(template)
    
    print(f"生成完成:")
    print(f"  - {OUTPUT_JSON}")
    print(f"  - {OUTPUT_HTML}")
    print(f"  分析周期: {start_date} ~ {end_date}")
    print(f"  近一周总收入: ¥{total_week1_income:,.0f}")
    print(f"  近一周总观看: {total_week1_views:,}")
    print(f"  收入周环比: {income_pct_str}")
    print(f"  观看周环比: {views_pct_str}")

if __name__ == "__main__":
    main()
